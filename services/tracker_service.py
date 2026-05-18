import os
import shutil

from datetime import datetime

from openpyxl import load_workbook


class TrackerService:

    @staticmethod
    def create_tracker_file(
        date_folder_path
    ):

        today = datetime.now().strftime(
            "%d_%m_%Y"
        )

        template_path = (
            "Template.xlsx"
        )

        report_name = (
            f"Report_{today}.xlsx"
        )

        report_path = os.path.join(
            date_folder_path,
            report_name
        )

        shutil.copy(
            template_path,
            report_path
        )

        return report_path

    @staticmethod
    def update_tracker(
            report_path,
            file_name,
            upload_status="",
            extraction_status=""
    ):

        workbook = load_workbook(
            report_path
        )

        sheet = workbook.active

        file_found = False

        for row in range(
                2,
                sheet.max_row + 1
        ):

            existing_file = sheet.cell(
                row=row,
                column=1
            ).value

            if existing_file == file_name:

                if upload_status:
                    sheet.cell(
                        row=row,
                        column=2
                    ).value = upload_status

                if extraction_status:
                    sheet.cell(
                        row=row,
                        column=3
                    ).value = extraction_status

                file_found = True

                break

        if not file_found:
            sheet.append([
                file_name,
                upload_status,
                extraction_status
            ])

        workbook.save(
            report_path
        )

        workbook.close()