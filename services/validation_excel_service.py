import os

from openpyxl import Workbook
from openpyxl import load_workbook


class ValidationExcelService:

    BASE_HEADERS = [

        "File Name",

        "File Path",

        "PO_Number",

        "PO_date",

        "Vendor_Id",

        "total_amount",

        "currency",

        "Vendor_Name",

        "Vendor_Email",

        "Ship_Address",

        "Supplier",

        "Buyer",

        "Bill_Address",

        "Payment_Terms",

        "Delivery_Terms"
    ]

    ITEM_FIELDS = [

        "material code",

        "description",

        "quantity",

        "unit_price",

        "Item_value"
    ]

    @staticmethod
    def create_excel_if_not_exists(
            excel_path,
            sheet_name
    ):

        if not os.path.exists(
                excel_path
        ):

            workbook = Workbook()

            default_sheet = workbook.active

            workbook.remove(
                default_sheet
            )

            sheet = workbook.create_sheet(
                title=sheet_name
            )

            for col_num, header in enumerate(
                    ValidationExcelService.BASE_HEADERS,
                    start=1
            ):

                sheet.cell(
                    row=1,
                    column=col_num
                ).value = header

            workbook.save(
                excel_path
            )

            workbook.close()

        else:

            workbook = load_workbook(
                excel_path
            )

            if sheet_name not in workbook.sheetnames:

                sheet = workbook.create_sheet(
                    title=sheet_name
                )

                for col_num, header in enumerate(
                        ValidationExcelService.BASE_HEADERS,
                        start=1
                ):

                    sheet.cell(
                        row=1,
                        column=col_num
                    ).value = header

                workbook.save(
                    excel_path
                )

            workbook.close()

    @staticmethod
    def ensure_item_headers(
            excel_path,
            sheet_name,
            item_count
    ):

        workbook = load_workbook(
            excel_path
        )

        sheet = workbook[
            sheet_name
        ]

        existing_headers = []

        for cell in sheet[1]:

            existing_headers.append(
                cell.value
            )

        next_column = (
            len(existing_headers) + 1
        )

        for item_index in range(
                item_count
        ):

            for field in (
                ValidationExcelService.ITEM_FIELDS
            ):

                header_name = (
                    f"item[{item_index}]."
                    f"{field}"
                )

                if (
                        header_name
                        not in existing_headers
                ):

                    sheet.cell(
                        row=1,
                        column=next_column
                    ).value = header_name

                    next_column += 1

        workbook.save(
            excel_path
        )

        workbook.close()

    @staticmethod
    def append_row(
            excel_path,
            sheet_name,
            file_name,
            file_path,
            extraction_data
    ):

        workbook = load_workbook(
            excel_path
        )

        sheet = workbook[
            sheet_name
        ]

        next_row = (
            sheet.max_row + 1
        )

        row_data = [

            file_name,

            file_path,

            extraction_data.get(
                "PO_Number"
            ),

            extraction_data.get(
                "PO_date"
            ),

            extraction_data.get(
                "Vendor_Id"
            ),

            extraction_data.get(
                "total_amount"
            ),

            extraction_data.get(
                "currency"
            ),

            extraction_data.get(
                "Vendor_Name"
            ),

            extraction_data.get(
                "Vendor_Email"
            ),

            extraction_data.get(
                "Ship_Address"
            ),

            extraction_data.get(
                "Supplier"
            ),

            extraction_data.get(
                "Buyer"
            ),

            extraction_data.get(
                "Bill_Address"
            ),

            extraction_data.get(
                "Payment_Terms"
            ),

            extraction_data.get(
                "Delivery_Terms"
            )
        ]

        items = extraction_data.get(
            "items",
            []
        )

        for item in items:

            row_data.extend([

                item.get(
                    "material code"
                ),

                item.get(
                    "description"
                ),

                item.get(
                    "quantity"
                ),

                item.get(
                    "unit_price"
                ),

                item.get(
                    "Item_value"
                )
            ])

        for col_num, value in enumerate(
                row_data,
                start=1
        ):

            sheet.cell(
                row=next_row,
                column=col_num
            ).value = value

        workbook.save(
            excel_path
        )

        workbook.close()